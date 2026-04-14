"""
LogiLink — Pipeline de Scoring d'Employabilité
================================================
Calcule un score /100 pour chaque étudiant en fonction de :
  - La matrice de couverture pondérée (matières × 24 métiers)
  - Le relevé de notes (avec gestion des cas L, LM, M)

Auteur : LogiLink 
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from pathlib import Path
import argparse
import json
import os
import re
import unicodedata
import warnings
from supabase import Client, create_client
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# 1. CONFIGURATION
# ─────────────────────────────────────────────

PARCOURS_PARAMS = {
    "L":  {"CMA": 1.00},
    "LM": {"CMA": 1.00},
    "M":  {"CMA": 1.00},
}

EPSILON = 1e-8  # évite division par zéro
SCORE_TABLE_NAME = os.getenv("EMPLOYABILITY_SCORE_TABLE", "score_employabilité").strip() or "score_employabilité"


def _load_env_file(env_path: Path) -> dict[str, str]:
    """
    Charge un fichier .env simple (KEY=VALUE).
    """
    if not env_path.exists():
        return {}

    values: dict[str, str] = {}
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key:
            values[key] = value

    return values


def _resolve_supabase_credentials() -> tuple[str, str]:
    """
    Résout les credentials Supabase depuis l'environnement ou des .env locaux.
    Priorité:
      1) Variables d'environnement du process
      2) employabilite_service/.env
      3) backend/.env
    """
    url = os.getenv("SUPABASE_URL", "").strip()
    key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip() or os.getenv("SUPABASE_KEY", "").strip()

    if url and key:
        return url, key

    base_dir = Path(__file__).resolve().parent
    candidate_env_files = [
        base_dir / ".env",
        base_dir.parent / "backend" / ".env",
    ]

    for env_file in candidate_env_files:
        env_values = _load_env_file(env_file)
        url = url or env_values.get("SUPABASE_URL", "").strip()
        key = key or env_values.get("SUPABASE_SERVICE_ROLE_KEY", "").strip() or env_values.get("SUPABASE_KEY", "").strip()
        if url and key:
            return url, key

    return url, key


def init_supabase_client() -> Client:
    """
    Initialise et retourne le client Supabase.
    """
    supabase_url, supabase_key = _resolve_supabase_credentials()
    if not supabase_url or not supabase_key:
        raise EnvironmentError(
            "SUPABASE_URL and SUPABASE_KEY must be set in environment variables or in employabilite_service/.env."
        )

    return create_client(supabase_url, supabase_key)


# ─────────────────────────────────────────────
# 2. CHARGEMENT DE LA MATRICE DE COUVERTURE
# ─────────────────────────────────────────────

def load_coverage_matrix(xlsx_path: str) -> pd.DataFrame:
    """
    Charge la matrice de couverture depuis le fichier Excel.
    Retourne un DataFrame indexé par 'matiere' avec 24 colonnes métier.
    """
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["Matrice de couverture"]
    rows = list(ws.iter_rows(values_only=True))

    # Trouver la ligne d'en-tête
    header_idx = next(
        i for i, r in enumerate(rows)
        if r[0] and "ilière" in str(r[0])
    )
    header = rows[header_idx]

    # Extraire les noms de métiers (colonnes 3 à 26)
    metier_names = []
    for cell in header[3:]:
        if cell:
            # Nettoyer le nom (retirer \n et code)
            name = str(cell).split("\n")[-1].strip()
            metier_names.append(name)

    # Parser les lignes de données
    records = []
    current_filiere = None
    current_semestre = None

    for row in rows[header_idx + 1:]:
        if not any(v is not None for v in row):
            continue

        # Propager filière et semestre (cellules fusionnées → None)
        if row[0] is not None:
            current_filiere = str(row[0]).strip()
        if row[1] is not None:
            current_semestre = str(row[1]).strip()

        matiere = row[2]
        if matiere is None:
            continue

        poids = list(row[3:3 + len(metier_names)])
        poids_clean = [float(p) if p is not None else 0.0 for p in poids]

        records.append({
            "filiere": current_filiere,
            "semestre": current_semestre,
            "matiere": str(matiere).strip(),
            **dict(zip(metier_names, poids_clean))
        })

    df = pd.DataFrame(records)
    return df


# ─────────────────────────────────────────────
# 3. CHARGEMENT DES NOTES
# ─────────────────────────────────────────────

def load_grades() -> pd.DataFrame:
    """
    Charge le relevé de notes.
    Normalise les notes sur [0, 1].
    """
    supabase = init_supabase_client()

    all_rows: list[dict] = []
    page_size = 1000
    start = 0
    response_count = None

    try:
        while True:
            response = (
                supabase
                .table("note")
                .select("*", count="exact")
                .range(start, start + page_size - 1)
                .execute()
            )

            if response_count is None:
                response_count = getattr(response, "count", None)

            batch = response.data or []
            if not batch:
                break

            all_rows.extend(batch)
            if len(batch) < page_size:
                break

            start += page_size
    except Exception as exc:
        raise RuntimeError(
            "Failed to fetch data from Supabase table 'note'."
        ) from exc

    df = pd.DataFrame(all_rows)

    _, resolved_key = _resolve_supabase_credentials()
    is_publishable_key = resolved_key.startswith("sb_publishable")
    response_count = getattr(response, "count", None)

    if (response_count == 0 or len(all_rows) == 0) and is_publishable_key:
        raise PermissionError(
            "Supabase returned 0 rows from table 'note' using a publishable key. "
            "This is usually caused by RLS policies blocking SELECT. "
            "Use SUPABASE_SERVICE_ROLE_KEY for backend jobs, or add a SELECT policy on public.note for your auth role."
        )

    if df.empty:
        return pd.DataFrame(
            columns=[
                "etudiant",
                "parcours_type",
                "filiere_licence",
                "filiere_master",
                "matiere",
                "moyenne_matiere",
                "note_norm",
            ]
        )

    df.columns = df.columns.str.strip()

    def normalize_column_name(name: str) -> str:
        normalized = unicodedata.normalize("NFKD", str(name))
        normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
        normalized = normalized.strip().lower()
        normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
        return normalized.strip("_")

    alias_map = {
        "etudiant": [
            "etudiant",
            "id_etudiant",
            "etudiant_id",
            "student_id",
            "id_student",
            "id",
        ],
        "parcours_type": [
            "parcours_type",
            "parcours",
            "type_parcours",
            "track_type",
        ],
        "filiere_licence": [
            "filiere_licence",
            "licence_filiere",
            "filiere_license",
            "license_filiere",
            "filiere_l",
            "filiere_lic",
        ],
        "filiere_master": [
            "filiere_master",
            "master_filiere",
            "filiere_m",
        ],
        "matiere": ["matiere", "matiere_nom", "nom_matiere", "subject", "module"],
        "moyenne_matiere": [
            "moyenne_matiere",
            "moyenne",
            "note",
            "grade",
            "average",
            "score",
        ],
    }

    normalized_to_original = {
        normalize_column_name(column): column for column in df.columns
    }
    rename_map: dict[str, str] = {}
    for target, aliases in alias_map.items():
        for alias in aliases:
            source_column = normalized_to_original.get(normalize_column_name(alias))
            if source_column and source_column not in rename_map:
                rename_map[source_column] = target
                break

    if rename_map:
        df = df.rename(columns=rename_map)

    required_columns = {"etudiant", "parcours_type", "matiere", "moyenne_matiere"}
    missing_columns = required_columns - set(df.columns)
    if missing_columns:
        raise ValueError(
            "Missing required columns from Supabase result after alias mapping: "
            f"{sorted(missing_columns)}. Available columns: {sorted(df.columns.tolist())}"
        )

    if "filiere_licence" not in df.columns:
        df["filiere_licence"] = pd.NA
    if "filiere_master" not in df.columns:
        df["filiere_master"] = pd.NA

    df["etudiant"] = df["etudiant"].astype(str).str.strip()
    df = df[df["etudiant"] != ""]

    df["parcours_type"] = df["parcours_type"].astype(str).str.strip().str.upper()

    df["note_norm"] = pd.to_numeric(df["moyenne_matiere"], errors="coerce") / 20.0
    df["matiere"] = df["matiere"].astype(str).str.strip()
    return df


# ─────────────────────────────────────────────
# 4. CALCUL DU SCORE PAR MÉTIER
# ─────────────────────────────────────────────

def score_per_metier(
    student_notes: pd.DataFrame,
    coverage_matrix: pd.DataFrame,
    metier_cols: list,
    parcours_type: str,
    filiere_licence: str | None = None,
    filiere_master: str | None = None
) -> pd.Series:
    """
    Calcule le RawScore[k] pour chaque métier k :
      RawScore[k] = Σ(C[m,k] * n[m]) / Σ(C[m,k])
    """
    # Filtrage filière: L -> Licence, M -> Master, LM -> Licence + Master
    filiere_mapping = {
        "Licence_Sciences_de_Transport": "Licence_en_Sciences_de_Transport",
        "Licence_Génie_Logistique": "Licence_Génie_Logistique",
        "Master_Recherche_STL": "Master_Recherche_STL",
        "Master_Pro_Génie_Industriel_et_Logistique": "Master_Pro_Génie_Industriel_et_Logistique",
    }
    normalized_licence = filiere_mapping.get(str(filiere_licence).strip()) if filiere_licence is not None else None
    normalized_master = filiere_mapping.get(str(filiere_master).strip()) if filiere_master is not None else None

    # Si des spécialisations sont fournies, privilégier un filtrage exact par filière.
    if normalized_licence is not None or normalized_master is not None:
        allowed_filieres = set()
        if normalized_licence is not None:
            allowed_filieres.add(normalized_licence)
        if normalized_master is not None:
            allowed_filieres.add(normalized_master)
        coverage_filtered = coverage_matrix[
            coverage_matrix["filiere"].astype(str).str.strip().isin(allowed_filieres)
        ].copy()
    else:
        # Comportement existant inchangé si aucune spécialisation n'est fournie.
        coverage_filtered = coverage_matrix.copy()

    if normalized_licence is None and normalized_master is None:
        # Filtrage historique par type de parcours
        parcours = str(parcours_type).strip().upper()
        if parcours == "L":
            coverage_filtered = coverage_matrix[
                coverage_matrix["filiere"].astype(str).str.contains("Licence", case=False, na=False)
            ].copy()
        elif parcours == "M":
            coverage_filtered = coverage_matrix[
                coverage_matrix["filiere"].astype(str).str.contains("Master", case=False, na=False)
            ].copy()
        elif parcours == "LM":
            coverage_filtered = coverage_matrix[
                coverage_matrix["filiere"].astype(str).str.contains("Licence|Master", case=False, na=False)
            ].copy()
        else:
            coverage_filtered = coverage_matrix.copy()

    # Préparer les matières (normalisation locale)
    notes_work = student_notes.copy()
    notes_work["matiere"] = notes_work["matiere"].astype(str).str.strip()

    # Cas LM: la colonne filiere_master sert aussi de tag au niveau ligne
    # (vide/NaN -> ligne Licence, rempli -> ligne Master).
    # On fait deux jointures séparées pour éviter les doubles appariements.
    parcours = str(parcours_type).strip().upper()
    if parcours == "LM" and (normalized_licence is not None or normalized_master is not None):
        licence_cov = coverage_filtered.iloc[0:0].copy()
        master_cov = coverage_filtered.iloc[0:0].copy()

        if normalized_licence is not None:
            licence_cov = coverage_filtered[
                coverage_filtered["filiere"].astype(str).str.strip() == normalized_licence
            ].copy()
        if normalized_master is not None:
            master_cov = coverage_filtered[
                coverage_filtered["filiere"].astype(str).str.strip() == normalized_master
            ].copy()

        filiere_master_tag = notes_work["filiere_master"]
        master_mask = filiere_master_tag.notna() & (filiere_master_tag.astype(str).str.strip() != "")

        licence_notes = notes_work[~master_mask].copy()
        master_notes = notes_work[master_mask].copy()

        merged_parts = []

        if not licence_notes.empty:
            matieres_licence = set(
                licence_cov["matiere"].dropna().astype(str).str.strip().tolist()
            )
            licence_notes = licence_notes[
                licence_notes["matiere"].isin(matieres_licence)
            ]
            merged_licence = licence_notes[["matiere", "note_norm"]].merge(
                licence_cov[["matiere"] + metier_cols],
                on="matiere",
                how="left"
            )
            merged_parts.append(merged_licence)

        if not master_notes.empty:
            matieres_master = set(
                master_cov["matiere"].dropna().astype(str).str.strip().tolist()
            )
            master_notes = master_notes[
                master_notes["matiere"].isin(matieres_master)
            ]
            merged_master = master_notes[["matiere", "note_norm"]].merge(
                master_cov[["matiere"] + metier_cols],
                on="matiere",
                how="left"
            )
            merged_parts.append(merged_master)

        if merged_parts:
            merged = pd.concat(merged_parts, ignore_index=True)
        else:
            merged = pd.DataFrame(columns=["matiere", "note_norm"] + metier_cols)
    else:
        # Comportement inchangé hors cas LM taggé
        matieres_filiere = set(
            coverage_filtered["matiere"].dropna().astype(str).str.strip().tolist()
        )
        notes_filtered = notes_work[
            notes_work["matiere"].isin(matieres_filiere)
        ]

        # Joindre les notes avec la matrice filtrée sur le nom de matière
        merged = notes_filtered[["matiere", "note_norm"]].merge(
            coverage_filtered[["matiere"] + metier_cols],
            on="matiere",
            how="left"
        )

    # Remplir les matières absentes de la matrice par 0
    merged[metier_cols] = merged[metier_cols].fillna(0.0)

    scores = {}
    for k in metier_cols:
        weights = merged[k].values
        notes = merged["note_norm"].values
        num = np.sum(weights * notes)
        den = np.sum(weights) + EPSILON
        scores[k] = num / den

    return pd.Series(scores)


# ─────────────────────────────────────────────
# 5. SCORE GLOBAL D'UN ÉTUDIANT
# ─────────────────────────────────────────────

def compute_student_score(
    student_id: str,
    grades_df: pd.DataFrame,
    coverage_matrix: pd.DataFrame,
    metier_cols: list
) -> dict:
    """
    Calcule le score final /100 pour un étudiant donné.
    Retourne un dict avec score global et scores par métier.
    """
    student_data = grades_df[grades_df["etudiant"] == student_id].copy()

    if student_data.empty:
        return {"etudiant": student_id, "score_final": None, "erreur": "Aucune donnée"}

    parcours = student_data["parcours_type"].iloc[0]
    params = PARCOURS_PARAMS.get(parcours, {"CMA": 1.0})

    def first_non_empty(series: pd.Series):
        values = series.dropna().astype(str).str.strip()
        values = values[values != ""]
        return values.iloc[0] if not values.empty else np.nan

    filiere_licence = first_non_empty(student_data["filiere_licence"])
    filiere_master = first_non_empty(student_data["filiere_master"])

    # Supprimer les lignes avec note manquante
    student_data = student_data.dropna(subset=["note_norm"])

    # Calcul des scores par métier
    raw_scores = score_per_metier(
        student_data,
        coverage_matrix,
        metier_cols,
        parcours,
        None if pd.isna(filiere_licence) else str(filiere_licence),
        None if pd.isna(filiere_master) else str(filiere_master),
    )

    # Score global brut (moyenne uniforme sur les 24 métiers)
    global_raw = raw_scores.mean()

    # Application CMA  + normalisation sur 100
    score_final = min(100, (global_raw * params["CMA"] ) * 100)

    result = {
        "etudiant": student_id,
        "parcours_type": parcours,
        "filiere_licence": filiere_licence,
        "filiere_master": filiere_master,
        "nb_matieres": len(student_data),
        "global_raw": round(global_raw, 4),
        "CMA": params["CMA"],
      
        "score_final": round(score_final, 2),
    }

    # Ajouter les scores par métier (arrondis à 4 décimales)
    for k, v in raw_scores.items():
        result[f"score_{k[:30]}"] = round(v, 4)

    return result


# ─────────────────────────────────────────────
# 6. PIPELINE PRINCIPAL
# ─────────────────────────────────────────────

def run_pipeline(xlsx_path: str) -> pd.DataFrame:
    """
    Lance le pipeline complet et retourne un DataFrame de résultats.
    """
    print("🔄 Chargement de la matrice de couverture...")
    coverage = load_coverage_matrix(xlsx_path)
    metier_cols = [c for c in coverage.columns if c not in ["filiere", "semestre", "matiere"]]
    print(f"   ✅ {len(coverage)} matières × {len(metier_cols)} métiers")

    print("🔄 Chargement des relevés de notes depuis Supabase...")
    grades = load_grades()
    student_ids = grades["etudiant"].unique()
    print(f"   ✅ {len(student_ids)} étudiants, {len(grades)} lignes")

    if len(student_ids) == 0:
        print("   ⚠️ Aucune ligne exploitable trouvée dans la table note.")
        return pd.DataFrame(), metier_cols

    print("⚙️  Calcul des scores...")
    results = []
    for sid in student_ids:
        results.append(compute_student_score(sid, grades, coverage, metier_cols))

    df_results = pd.DataFrame(results)
    df_results = df_results.sort_values("etudiant").reset_index(drop=True)

    return df_results, metier_cols


def compute_single_student_score(
    student_id: str,
    xlsx_path: str,
) -> dict:
    """
    Exécute la logique existante de score pour un seul étudiant.
    """
    coverage = load_coverage_matrix(xlsx_path)
    metier_cols = [c for c in coverage.columns if c not in ["filiere", "semestre", "matiere"]]
    grades = load_grades()

    student_key = str(student_id).strip()
    student_rows = grades[grades["etudiant"].astype(str).str.strip() == student_key]

    if student_rows.empty:
        raise ValueError(f"No note rows found for student '{student_key}'")

    return compute_student_score(student_key, grades, coverage, metier_cols)


def _to_python_scalar(value):
    if isinstance(value, np.generic):
        return value.item()
    if pd.isna(value):
        return None
    return value


def _extract_missing_column_from_error(exc: Exception) -> str | None:
    message = str(exc)
    match = re.search(r"Could not find the '([^']+)' column", message)
    if match:
        return match.group(1)
    return None


def _error_mentions_on_conflict_constraint(exc: Exception) -> bool:
    message = str(exc).lower()
    return "no unique or exclusion constraint matching the on conflict specification" in message


def _upsert_with_column_pruning(supabase: Client, table_name: str, row: dict, student_key: str) -> None:
    payload = dict(row)

    def _prune_column_if_needed(exc: Exception) -> bool:
        missing_column = _extract_missing_column_from_error(exc)
        if not missing_column or missing_column == "etudiant" or missing_column not in payload:
            return False
        payload.pop(missing_column, None)
        return True

    for _ in range(20):
        try:
            supabase.table(table_name).upsert(payload, on_conflict="etudiant").execute()
            return
        except Exception as exc:
            if _prune_column_if_needed(exc):
                continue

            if not _error_mentions_on_conflict_constraint(exc):
                raise

            # Fallback when etudiant is not declared UNIQUE in DB schema:
            # perform read-then-update/insert while still pruning unknown columns.
            lookup = (
                supabase
                .table(table_name)
                .select("etudiant")
                .eq("etudiant", student_key)
                .limit(1)
                .execute()
            )
            exists = bool(lookup.data)

            for _ in range(20):
                try:
                    if exists:
                        supabase.table(table_name).update(payload).eq("etudiant", student_key).execute()
                    else:
                        supabase.table(table_name).insert(payload).execute()
                    return
                except Exception as nested_exc:
                    if _prune_column_if_needed(nested_exc):
                        continue
                    raise

    raise RuntimeError(
        f"Failed to upsert into Supabase table '{table_name}' after pruning unknown columns."
    )


def upsert_student_score_to_supabase(result: dict, table_name: str = SCORE_TABLE_NAME) -> None:
    """
    Met à jour/insère le score d'un étudiant dans Supabase.
    """
    supabase = init_supabase_client()

    row = {str(k): _to_python_scalar(v) for k, v in result.items()}
    student_key = str(row.get("etudiant", "")).strip()
    if not student_key:
        raise ValueError("Cannot upsert score: missing etudiant in result payload")

    row["etudiant"] = student_key
    _upsert_with_column_pruning(supabase, table_name, row, student_key)


def upsert_student_score_to_csv(result: dict, output_csv_path: str) -> None:
    """
    Compatibilité rétroactive: persiste désormais le score en Supabase.
    """
    _ = output_csv_path
    upsert_student_score_to_supabase(result)


def run_single_student_pipeline(
    student_id: str,
    xlsx_path: str,
    output_csv_path: str,
) -> dict:
    """
    Calcule le score d'un étudiant connecté et met à jour le CSV.
    """
    result = compute_single_student_score(student_id, xlsx_path)
    upsert_student_score_to_csv(result, output_csv_path)
    return result


# ─────────────────────────────────────────────
# 7. ANALYSE & VALIDATION
# ─────────────────────────────────────────────




def show_sample(df: pd.DataFrame, n: int = 5) -> None:
    """Affiche un échantillon de résultats."""
    cols = ["etudiant", "parcours_type", "filiere_licence", "filiere_master",
            "nb_matieres", "global_raw", "score_final"]
    print(f"\n📋 Échantillon de {n} résultats par parcours :")
    for pt in ["L", "M", "LM"]:
        subset = df[df["parcours_type"] == pt].head(n)[cols]
        print(f"\n  Parcours {pt}:")
        print(subset.to_string(index=False))


# ─────────────────────────────────────────────
# 8. POINT D'ENTRÉE
# ─────────────────────────────────────────────

if __name__ == "__main__":
    base_dir = Path(__file__).resolve().parent

    parser = argparse.ArgumentParser(description="LogiLink employability scoring pipeline")
    parser.add_argument("--student-id", dest="student_id", help="Compute and upsert score only for this student")
    parser.add_argument("--xlsx-path", dest="xlsx_path", default=str(base_dir / "matrice_coverture.xlsx"))
    parser.add_argument("--output-path", dest="output_path", default=str(base_dir / "logilink_scores.csv"))
    parser.add_argument("--json", dest="json_output", action="store_true", help="Print machine-readable JSON output")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx_path)
    output_path = Path(args.output_path)

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Fichier introuvable: {xlsx_path}")

    if args.student_id:
        try:
            result = run_single_student_pipeline(str(args.student_id), str(xlsx_path), str(output_path))
            payload = {
                "ok": True,
                "studentId": str(result.get("etudiant", "")).strip(),
                "scoreFinal": result.get("score_final"),
                "outputPath": f"supabase:{SCORE_TABLE_NAME}",
            }
            if args.json_output:
                print(f"RESULT_JSON:{json.dumps(payload, ensure_ascii=False)}")
            else:
                print(f"✅ Score employabilité mis à jour pour étudiant {payload['studentId']} → {payload['outputPath']}")
        except Exception as exc:
            err_payload = {
                "ok": False,
                "studentId": str(args.student_id).strip(),
                "error": str(exc),
            }
            if args.json_output:
                print(f"RESULT_JSON:{json.dumps(err_payload, ensure_ascii=False)}")
            raise
        raise SystemExit(0)

    df_scores, metier_cols = run_pipeline(str(xlsx_path))

    if df_scores.empty:
        print("\n⚠️ Aucun score généré: vérifiez les colonnes et les données de la table note.")
        raise SystemExit(0)

    show_sample(df_scores, n=3)
    df_scores.to_csv(str(output_path), index=False, encoding="utf-8-sig")
    print(f"\n✅ Résultats sauvegardés → {output_path}")
    print(f"   Colonnes : {list(df_scores.columns[:10])} ...")
